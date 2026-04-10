"""
app/agents/designer.py — Agent 4: Brand & Format Specialist
Skill: Calls deterministic document builders directly — no LLM code generation.

Pipeline:
  PPTX → app/tools/document_skills/pptx_builder.py
  DOCX → app/tools/document_skills/docx_builder.py
  XLSX → app/tools/document_skills/xlsx_builder.py (openpyxl)
  PDF  → reportlab fallback
"""

from pathlib import Path
from app.agents.state import AgentState
from app.agents.messenger import post_message
from app.core.config import settings
from app.core.logger import setup_logger

log = setup_logger("designer")

OUTPUT_DIR = Path(settings.OUTPUT_DIR)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def run_designer(state: AgentState) -> AgentState:
    draft_text    = state.get("draft_text", "")
    output_format = state.get("output_format", "docx")
    session_id    = state.get("session_id", "output")
    output_path   = str(OUTPUT_DIR / f"{session_id}.{output_format}")

    log.info("Designer: Starting brand formatting phase.")
    log.debug(f"Designer: format={output_format} | output={output_path} | draft_chars={len(draft_text)}")
    job_id = state.get("job_id", "")
    post_message(job_id, "designer", f"🎨 On it @synthesizer. Building {output_format.upper()} with brand colors (#1A1A2E, #E94560)...")

    tdd_passed = False
    exec_error = ""

    try:
        if output_format == "pptx":
            from app.tools.document_skills.pptx_builder import build_pptx
            tdd_passed = build_pptx(draft_text, output_path)

        elif output_format == "docx":
            from app.tools.document_skills.docx_builder import build_docx
            tdd_passed = build_docx(draft_text, output_path)

        elif output_format == "xlsx":
            tdd_passed = _build_xlsx(draft_text, output_path)

        elif output_format == "pdf":
            tdd_passed = _build_pdf(draft_text, output_path)

        else:
            raise ValueError(f"Unsupported output format: {output_format}")

        if tdd_passed:
            log.info(f"Designer: Document built → {output_path}")
            log.debug(f"Designer: File size = {Path(output_path).stat().st_size} bytes")
            post_message(job_id, "designer", f"✅ Document built → `{Path(output_path).name}` ({Path(output_path).stat().st_size // 1024} KB). @inspector please review.", "success")
        else:
            exec_error = f"Builder returned False for {output_format}"
            log.error(f"Designer: {exec_error}")

    except Exception as e:
        exec_error = str(e)
        log.error(f"Designer: Build failed — {exec_error}")
        post_message(job_id, "designer", f"❌ Build failed: {exec_error[:200]}", "error")
        tdd_passed = False

    return {
        **state,
        "formatted_file_path": output_path if tdd_passed else "",
        "formatting_code": f"# Used {output_format}_builder skill directly",
        "tdd_passed": tdd_passed,
        "error_log": exec_error if not tdd_passed else state.get("error_log", ""),
    }


def _build_xlsx(draft_text: str, output_path: str) -> bool:
    """Build a structured Excel file from draft text sections.

    Handles three content patterns the Synthesizer produces:
      1. Markdown pipe tables  → proper multi-column rows
      2. **Key:** Value pairs  → two-column key/value rows
      3. Bullet / plain lines  → single-column content rows
    All markdown artifacts (**bold**, *italic*) are stripped from cell values.
    """
    import re
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Brand styles ──────────────────────────────────────────────────
    NAVY        = "1A1A2E"
    CRIMSON     = "E94560"
    OFF_WHITE   = "EAEAEA"
    STRIPE_EVEN = "F5F5F8"
    STRIPE_ODD  = "FFFFFF"

    SECTION_FONT  = Font(name="Calibri", bold=True, color=OFF_WHITE, size=12)
    SECTION_FILL  = PatternFill("solid", fgColor=NAVY)
    COL_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    COL_HDR_FILL  = PatternFill("solid", fgColor=CRIMSON)
    KEY_FONT      = Font(name="Calibri", bold=True, color=NAVY, size=11)
    BODY_FONT     = Font(name="Calibri", size=11, color="333333")
    THIN_SIDE     = Side(style="thin", color="DDDDDD")
    THIN_BORDER   = Border(bottom=THIN_SIDE)

    def _clean(text: str) -> str:
        """Strip markdown bold/italic markers and excess whitespace."""
        text = re.sub(r"\*{1,3}(.*?)\*{1,3}", r"\1", text)
        text = re.sub(r"`(.*?)`", r"\1", text)
        return text.strip()

    def _is_separator(line: str) -> bool:
        """Detect markdown table separator rows like |---|---|"""
        return bool(re.match(r"^\|[\s\-|:]+\|$", line.strip()))

    def _parse_pipe_row(line: str) -> list[str]:
        """Split a markdown pipe-table row into cleaned cells."""
        cells = [_clean(c) for c in line.strip().strip("|").split("|")]
        return [c for c in cells if c]  # drop empty edge cells

    def _write_row(ws, row_idx: int, cells: list, font, fill=None,
                   align_center=False, row_height=None):
        for ci, val in enumerate(cells, start=1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.font = font
            c.border = THIN_BORDER
            c.alignment = Alignment(
                horizontal="center" if align_center else "left",
                vertical="center",
                wrap_text=True,
            )
            if fill:
                c.fill = fill
        if row_height:
            ws.row_dimensions[row_idx].height = row_height

    # ── Parse sections ────────────────────────────────────────────────
    sections = re.split(r"^##\s+", draft_text, flags=re.MULTILINE)
    row = 1

    for section in sections:
        if not section.strip():
            continue
        lines = section.strip().splitlines()
        heading = _clean(lines[0])
        body_lines = [l for l in lines[1:] if l.strip()]

        # ── Section heading spanning 6 columns ───────────────────────
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=heading)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 24
        row += 1

        if not body_lines:
            row += 1
            continue

        # ── Detect content type ───────────────────────────────────────
        pipe_lines = [l for l in body_lines if l.strip().startswith("|")]
        kv_lines   = [l for l in body_lines if re.match(r"^\*{0,2}\w[^|]*:\*{0,2}\s+\S", l.strip())]

        if len(pipe_lines) >= 2:
            # ── Markdown pipe table ───────────────────────────────────
            header_written = False
            for line in body_lines:
                line = line.strip()
                if not line or _is_separator(line):
                    continue
                cells = _parse_pipe_row(line)
                if not cells:
                    continue
                if not header_written:
                    _write_row(ws, row, cells, COL_HDR_FONT, COL_HDR_FILL,
                               align_center=True, row_height=20)
                    header_written = True
                else:
                    stripe = STRIPE_EVEN if row % 2 == 0 else STRIPE_ODD
                    _write_row(ws, row, cells, BODY_FONT,
                               PatternFill("solid", fgColor=stripe.lstrip("#")),
                               row_height=18)
                row += 1

        elif len(kv_lines) > len(body_lines) // 2:
            # ── Key: Value pairs → two-column layout ─────────────────
            # Write a small header
            _write_row(ws, row, ["Property", "Value"], COL_HDR_FONT, COL_HDR_FILL,
                       align_center=True, row_height=20)
            row += 1
            for line in body_lines:
                line = line.strip().lstrip("•-* ")
                if not line:
                    continue
                # Split on first colon
                match = re.match(r"^(.*?):\s+(.+)$", _clean(line))
                if match:
                    key, val = match.group(1).strip(), match.group(2).strip()
                else:
                    key, val = _clean(line), ""
                stripe = STRIPE_EVEN if row % 2 == 0 else STRIPE_ODD
                fill = PatternFill("solid", fgColor=stripe.lstrip("#"))
                kc = ws.cell(row=row, column=1, value=key)
                kc.font = KEY_FONT
                kc.fill = fill
                kc.border = THIN_BORDER
                kc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                vc = ws.cell(row=row, column=2, value=val)
                vc.font = BODY_FONT
                vc.fill = fill
                vc.border = THIN_BORDER
                vc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[row].height = 18
                row += 1

        else:
            # ── Bullet / plain lines → single-column content ──────────
            for line in body_lines:
                line = _clean(line.strip().lstrip("•-* "))
                if not line:
                    continue
                # Try pipe split as a last resort
                if "|" in line:
                    parts = [p.strip() for p in line.split("|") if p.strip()]
                else:
                    parts = [line]
                stripe = STRIPE_EVEN if row % 2 == 0 else STRIPE_ODD
                _write_row(ws, row, parts, BODY_FONT,
                           PatternFill("solid", fgColor=stripe.lstrip("#")),
                           row_height=18)
                row += 1

        row += 1  # blank gap between sections

    # ── Auto-size columns ─────────────────────────────────────────────
    col_widths: dict[int, int] = {}
    for r in ws.iter_rows():
        for c in r:
            if c.value:
                col_widths[c.column] = max(
                    col_widths.get(c.column, 8),
                    min(len(str(c.value)), 60)
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 4

    # ── Footer ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=6)
    footer = ws.cell(row=row + 1, column=1, value="CONFIDENTIAL — AgentPress Internal")
    footer.font = Font(name="Calibri", size=9, italic=True, color="999999")
    footer.alignment = Alignment(horizontal="center")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    assert Path(output_path).exists()
    return True


def _build_pdf(draft_text: str, output_path: str) -> bool:
    """Build a basic PDF from draft text using reportlab."""
    import re
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

    BRAND_NAVY  = colors.HexColor("#1A1A2E")
    BRAND_RED   = colors.HexColor("#E94560")

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=3*cm, rightMargin=3*cm,
                            topMargin=2.5*cm, bottomMargin=2.5*cm)
    styles = getSampleStyleSheet()
    heading_style = ParagraphStyle("BrandH1", parent=styles["Heading1"],
                                   textColor=BRAND_RED, fontName="Helvetica-Bold", fontSize=14)
    body_style    = ParagraphStyle("BrandBody", parent=styles["Normal"],
                                   textColor=BRAND_NAVY, fontName="Helvetica", fontSize=11,
                                   leading=16)

    story = []
    sections = re.split(r"^##\s+", draft_text, flags=re.MULTILINE)

    for section in sections:
        if not section.strip():
            continue
        lines = section.strip().splitlines()
        story.append(Paragraph(lines[0].strip(), heading_style))
        story.append(Spacer(1, 0.3*cm))
        for line in lines[1:]:
            line = line.strip()
            if line:
                story.append(Paragraph(line.lstrip("•-* "), body_style))
        story.append(Spacer(1, 0.5*cm))

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    doc.build(story)
    assert Path(output_path).exists()
    return True
