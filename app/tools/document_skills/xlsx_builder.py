"""
app/tools/document_skills/xlsx_builder.py

Builds a branded multi-sheet Excel workbook from pipe-delimited table data.

Expected input format from the Synthesizer:

    ## Sheet Title
    Col A | Col B | Col C | Col D
    val1  | val2  | val3  | val4
    val5  | val6  | val7  | val8

Each ## heading becomes a separate worksheet tab.
The first data row after the heading is treated as the column header row.
"""

import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter


# ── Brand tokens ──────────────────────────────────────────────────────────────
NAVY        = "1A1A2E"
CRIMSON     = "E94560"
OFF_WHITE   = "EAEAEA"
STRIPE_A    = "F7F7FA"   # even rows
STRIPE_B    = "FFFFFF"   # odd rows
FOOTER_GREY = "999999"

THIN = Side(style="thin", color="DDDDDD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOTTOM_ONLY = Border(bottom=Side(style="thin", color="CCCCCC"))


def _clean(text: str) -> str:
    """Strip markdown bold/italic/code markers."""
    text = re.sub(r"\*{1,3}(.*?)\*{1,3}", r"\1", text)
    text = re.sub(r"`(.*?)`", r"\1", text)
    return text.strip()


def _is_separator(line: str) -> bool:
    return bool(re.match(r"^[\|\s\-:]+$", line.strip()))


def _split_pipe_row(line: str) -> list[str]:
    """Split a pipe-delimited row into cleaned cell values."""
    return [_clean(c) for c in line.strip().strip("|").split("|")]


def _style_header_row(ws, row_idx: int, num_cols: int):
    """Apply crimson header styling to a row."""
    for ci in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=ci)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=CRIMSON)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 22


def _style_data_row(ws, row_idx: int, num_cols: int, stripe: bool):
    """Apply zebra-stripe styling to a data row."""
    bg = STRIPE_A if stripe else STRIPE_B
    for ci in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=ci)
        is_first_col = ci == 1
        cell.font = Font(
            name="Calibri",
            bold=is_first_col,
            size=11,
            color=NAVY if is_first_col else "333333",
        )
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 18


def _style_section_row(ws, row_idx: int, num_cols: int, text: str):
    """Apply navy section-heading style spanning all columns."""
    ws.merge_cells(
        start_row=row_idx, start_column=1,
        end_row=row_idx, end_column=max(num_cols, 1)
    )
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font = Font(name="Calibri", bold=True, color=OFF_WHITE, size=12)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 24


def _auto_size_columns(ws):
    """Set column widths based on max content length."""
    col_widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and not cell.data_type == "n":
                col_widths[cell.column] = max(
                    col_widths.get(cell.column, 8),
                    min(len(str(cell.value)), 55),
                )
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 4


def _parse_sections(draft_text: str) -> list[dict]:
    """
    Split draft_text into sections by ## headings.
    Returns list of {title, lines} dicts.
    """
    raw_sections = re.split(r"^##\s+", draft_text, flags=re.MULTILINE)
    sections = []
    for block in raw_sections:
        block = block.strip()
        if not block:
            continue
        lines = block.splitlines()
        title = _clean(lines[0])
        body = [l for l in lines[1:] if l.strip() and not _is_separator(l)]
        sections.append({"title": title, "lines": body})
    return sections


def _write_sheet(wb: Workbook, section: dict, sheet_idx: int) -> None:
    """Write one section as a worksheet."""
    title = section["title"][:31]  # Excel tab name limit
    lines = section["lines"]

    # Create or reuse sheet
    if sheet_idx == 0:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title=title)

    if not lines:
        return

    # Freeze top rows
    ws.freeze_panes = "A3"

    # ── Detect if content is pipe-table or key-value or plain ────────
    pipe_lines = [l for l in lines if "|" in l]
    kv_lines   = [l for l in lines if re.match(r"^[^|]+:\s+\S", _clean(l.lstrip("•-* ")))]

    row = 1

    if len(pipe_lines) >= 2:
        # ── Pipe table mode ──────────────────────────────────────────
        rows_data = [_split_pipe_row(l) for l in pipe_lines]
        num_cols = max(len(r) for r in rows_data)

        # Sheet title banner
        _style_section_row(ws, row, num_cols, section["title"])
        row += 1

        # Column headers (first pipe row)
        for ci, val in enumerate(rows_data[0], start=1):
            ws.cell(row=row, column=ci, value=val)
        _style_header_row(ws, row, num_cols)
        row += 1

        # Data rows
        for di, data_row in enumerate(rows_data[1:]):
            for ci, val in enumerate(data_row, start=1):
                ws.cell(row=row, column=ci, value=val)
            _style_data_row(ws, row, num_cols, stripe=di % 2 == 0)
            row += 1

    elif len(kv_lines) > len(lines) // 2:
        # ── Key: Value mode → 2-column layout ───────────────────────
        num_cols = 2
        _style_section_row(ws, row, num_cols, section["title"])
        row += 1

        # Header
        ws.cell(row=row, column=1, value="Property")
        ws.cell(row=row, column=2, value="Value")
        _style_header_row(ws, row, num_cols)
        row += 1

        for di, line in enumerate(lines):
            line = _clean(line.strip().lstrip("•-* "))
            if not line:
                continue
            match = re.match(r"^(.*?):\s+(.+)$", line)
            key = match.group(1).strip() if match else line
            val = match.group(2).strip() if match else ""
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=val)
            _style_data_row(ws, row, num_cols, stripe=di % 2 == 0)
            row += 1

    else:
        # ── Plain lines → single-column list ────────────────────────
        num_cols = 1
        _style_section_row(ws, row, num_cols, section["title"])
        row += 1

        for di, line in enumerate(lines):
            line = _clean(line.strip().lstrip("•-* "))
            if not line:
                continue
            ws.cell(row=row, column=1, value=line)
            _style_data_row(ws, row, num_cols, stripe=di % 2 == 0)
            row += 1

    # Footer
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(num_cols, 1))
    footer = ws.cell(row=row, column=1, value="CONFIDENTIAL — AgentPress Internal")
    footer.font = Font(name="Calibri", size=9, italic=True, color=FOOTER_GREY)
    footer.alignment = Alignment(horizontal="center")

    _auto_size_columns(ws)


def build_xlsx(draft_text: str, output_path: str) -> bool:
    """
    Build a branded multi-sheet Excel workbook from pipe-delimited draft text.

    Args:
        draft_text: Synthesizer output with ## headings and pipe-table rows.
        output_path: Absolute path to write the .xlsx file.

    Returns:
        True if the file was created successfully.
    """
    wb = Workbook()
    sections = _parse_sections(draft_text)

    if not sections:
        # Fallback: single sheet with raw text
        ws = wb.active
        ws.title = "Sheet1"
        for i, line in enumerate(draft_text.splitlines(), start=1):
            if line.strip():
                ws.cell(row=i, column=1, value=_clean(line))
    else:
        for idx, section in enumerate(sections):
            _write_sheet(wb, section, idx)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    assert Path(output_path).exists()
    return True
